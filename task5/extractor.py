#-*- coding: utf-8 -*-

from collections import namedtuple
import openpyxl
from transliterate import translit
import random

#-----------------------------------------------------------------------------
# On the DATE HOME_TEAM welcomed GUEST_TEAM on the ice of CITY's ARENA.
# HOME_TEAM met GUEST_TEAM on ARENA rink in CITY on DATE. 
def generate_intro_phrase(date, 
                          home_team_name, 
                          guest_team_name, 
                          arena_name, 
                          city):
    selector = random.randint(0, 1)
    result = ""
    if selector == 0:
        result += "On " + date + " " + home_team_name + " welcomed " 
        result += guest_team_name + " on the ice of " + arena_name + " in "
        result += city + "."
    if selector == 1:
        result += home_team_name + " met " + guest_team_name + " on "
        result += arena_name + " rink in " + city + " on " + date + "."
    return result 

#-----------------------------------------------------------------------------
def generate_first_goal_phrase(team, actor_1, minute, actor_2, actor_3):
    selector = random.randint(0, 4)
    result = ""
    if selector == 0:
        result += team + " was the first to score through " + actor_1
        result += " in the " + add_ending(minute) + " " + "minute."
    if selector == 1:
        result += "It was " + team + " that struck first, with the effort of "
        result += actor_1 + " in the " + add_ending(minute) + " minute."
    if selector == 2:
        result += team + " opened the scoring through " + actor_1 + " in the "
        result += add_ending(minute) + " minute."
    if selector == 3:
        result += actor_1 + "'s goal came first in the " + add_ending(minute)
        result += " minute."
    if selector == 4:
        result += actor_1 + " of " + team + " was the first to score, " 
        result += str(minute)
        result += " minutes into the game"
        if not actor_2 is None:
            result += ", with the assistance of "
            result += actor_2
            if not actor_3 is None:
                result += " and " + actor_3
        result += "."
    return result
    
#-----------------------------------------------------------------------------    
def generate_final_goal_phrase(team, actor_1, minute, period):
    result = ""
    selector = random.randint(0, 2)
    if selector == 0:
        result += "In the " + add_ending(minute) + " minute, " + actor_1
        result += " completed the scoring."
    if selector == 1:
        result += team + " closed the game with " + actor_1
        result += "'s goal during the " + period + " period."
    if selector == 2:
        result += "The winner came in the " + add_ending(minute) + " minute."
    return result

#-----------------------------------------------------------------------------
def generate_penalty_phrase(actor, minute):
    result = ""
    selector = random.randint(0, 1)
    if selector == 0:
        result += "In the " + add_ending(minute) + " minute, " + actor
        result += " went to penalty box."
    if selector == 1:
        result += "In the " + add_ending(minute) + " minute, " + actor
        result += " received a penalty."
    return result


#-----------------------------------------------------------------------------
# Score_home_team and score_guest_team are computing after goal. 
def generate_goals_phrase(team, 
                          actor_1, 
                          minute, 
                          score_home_team, 
                          score_guest_team, 
                          team_is_home_team):
    selector = random.randint(0, 5)
    result = ""
    if selector == 0:
        result += "Then " + team + " scored through "
        result += actor_1 + "."
    if selector == 1:
        result += "After that " + actor_1 + " made it " + score_home_team
        result += "-" + score_guest_team
        result += " in the " + add_ending(minute) + " minute."
    if selector == 2:
        prev_score_home = int(score_home_team) - int(team_is_home_team)
        prev_score_guest = int(score_guest_team) - int(not team_is_home_team)
        if (prev_score_home == prev_score_guest) or \
            (prev_score_home > prev_score_guest and team_is_home_team) or \
            (prev_score_guest > prev_score_home and not team_is_home_team):
            result += "After that " + actor_1 + " scored to put his team "
            result += score_home_team + "-" + score_guest_team + " up."
        else:
            return generate_goals_phrase(team, actor_1, minute, 
                                         score_home_team, 
                                         score_guest_team, 
                                         team_is_home_team)
             
    if selector == 3:
        result += "Later " + actor_1 + " of " + team + " took it to "
        result += score_home_team + "-" + score_guest_team + " in the "
        result += add_ending(minute) + " minute."
    if selector == 4:
        result += "Then " + team + " went to " + score_home_team + "-"
        result += score_guest_team
        result += " thanks to " + actor_1 + "'s goal."
    if selector == 5:
        result += "The goal came in the " + add_ending(minute) + " minute, " 
        result += actor_1 + " being the architect."
    return result

#-----------------------------------------------------------------------------
def generate_injury_phrase(actor, minute):
    result = ""
    selector = random.randint(0, 2)
    if selector == 0:
        result += actor + " received an injury and was removed from the rink."
    if selector == 1:
        result += "In the " + add_ending(minute) + " minute " + actor
        result += " was injured."
    if selector == 2:
        result += actor + "'s injury left him out of the game."
    return result
    
#-----------------------------------------------------------------------------
def generate_single_goal_phrase(actor, minute):
    result = ""
    result += "The outcome was determined by a single goal in the "
    result += add_ending(minute) + " minute, fired by " + actor
    result += "."
    return result

#-----------------------------------------------------------------------------    
def generate_penalty_shot_phrase(team, minute, outcome):
    result = ""
    result += "In the " + add_ending(minute) + " " + team 
    result += " was assigned a penalty shot, which was " + outcome
    result += "."
    return result    
    
#-----------------------------------------------------------------------------
# period - int, action - bool, scores -- int
def generate_comment_after_period(action, period, 
                                  home_score, guest_score, 
                                  home_team, guest_team):
    result = ""
    home_score = int(home_score) 
    guest_score = int(guest_score)
    if period == 1:
        if action:
            result += "First period ended with the score of "
            result += str(home_score) + "-" + str(guest_score) + "."
        else:
            result += "First period was not sparked by any action."
        if home_score > guest_score:
            result += home_team + " was in the lead."
        elif home_score < guest_score:
            result += guest_team + " was in the lead."
        else:
            result += "\nThe tie was to be broken in the next period."
    if period == 2:
        if action:
            result += "The score after the second period was "
            result += str(home_score) + "-" + str(guest_score) + "."
        else:
            result += "Second period did not bring anything to the score."
        if home_score > guest_score:
            result += "\n" + home_team + " was in the lead."
        elif home_score < guest_score:
            result += "\n" + guest_team + " was in the lead."
        else:
            result += "\nSo, the score was leveled by the end of the period."
    if period == 3:
        if action:
            result += "The resulting score was "
            result += str(home_score) + "-" + str(guest_score) + "."
        else:
            result += "The last period did not change the score."
        if home_score > guest_score:
            result += "\nThe " + home_team + " beat " + guest_team + "."
        elif home_score < guest_score:
            result += "\nThe " + guest_team + " beat " + home_team + "."
        else:
            result += "\nThere was a tie, so the game went on to the" 
            result += " overtime."    
    return result
    
#-----------------------------------------------------------------------------
def add_ending(number):
    if type(number) == int:
        return add_ending(str(number))
    
    if type(number) == str or type(number) == unicode:
        if number == '11':
            return "11th"
        if number == '12':
            return "12th"
        
        last_digit = int(number[-1])
        if last_digit == 1:
            return number + "st"
        if last_digit == 2:
            return number +  "nd"
        if last_digit == 3:
            return number +  "rd"
        return number +  "th"        

#-----------------------------------------------------------------------------
def generate_overtime_ending_phrase(actor, team):
    result = "The tie was resolved in the overtime by " + actor +" of "
    result += team + "."
    return result

#-----------------------------------------------------------------------------
def generate_after_penalties_phrase(actor, team, goals, attempts):
    if goals > 1:
        result = str(goals) + " goals were scored (from "
    else: 
        result = str(goals) + " goal was scored (from "
    result += str(attempts) + " attempts),"
    result += " and the shootout ended with a spectacular goal by "
    result += actor + " of " + team + "."
    return result 
#-----------------------------------------------------------------------------
def generate_end_phrase(home_score, guest_score):
    result = "The match ends with the score " + str(home_score) 
    result += "-" + str(guest_score) + "."  
    return result

#-----------------------------------------------------------------------------
def swapper(name):    
    space_position = name.find(" ")
    
    name = name[space_position:] + " " + name[:space_position]
    name = name.rstrip().lstrip()
    
    return name

#-----------------------------------------------------------------------------
if __name__ == "__main__":
    # Load data.
    home_team = "home-team"
    guest_team = "guest-team"
    
    is_first_goal_scored = False 
    
    Player = namedtuple('Player', 
                        ("number", "name", "position"), 
                        verbose = False);

    wb = openpyxl.load_workbook('hockey.xlsx')
    ws = wb[wb.get_sheet_names()[0]]
    
    teams = {ws["E2"].value : home_team,
             ws["F2"].value : guest_team}
    
    team_names = {home_team : ws["E2"].value,
                  guest_team: ws["F2"].value}
    
    scores = {home_team :  0,
              guest_team : 0}
    
    # Generate intro section.
    match_date = ws["A2"].value
    intro = generate_intro_phrase(
                           str(match_date.strftime("%B")) + " " + \
                           str(match_date.strftime("%d")), 
                           team_names[home_team], 
                           team_names[guest_team], 
                           translit(ws["B2"].value, "ru", 
                                            reversed=True), 
                           translit(ws["C2"].value, "ru", 
                                            reversed=True))
    print intro
    # End of intro section.
    
    last_row = ws.get_highest_row()
    home_final_score = ws["B" + str(last_row - 1)].value
    guest_final_score = ws["B" + str(last_row)].value
    
    
    players = {home_team : {},
               guest_team : {}}
    
    # First line is a name of team.
    home_team_players_start = 4
    home_team_players_end = 26
    
    guest_team_players_start = 28
    guest_team_players_end = 50
    
    # If failed here, change previous variables.
    assert(ws["A" + str(home_team_players_start)].value == home_team)
    assert(ws["A" + str(guest_team_players_start)].value == guest_team)
    
    
    for i in xrange(home_team_players_start + 1, home_team_players_end + 1):
        player_number = ws["B" + str(i)].value
        player_name = ws["C" + str(i)].value
        player_position = ws["D" + str(i)].value
        player_name = translit(player_name, "ru", reversed=True)
        players[home_team][player_number] = Player(player_number, 
                                                   swapper(player_name), 
                                                   player_position)
    
    for i in xrange(guest_team_players_start + 1, guest_team_players_end + 1):
        player_number = ws["B" + str(i)].value
        player_name = ws["C" + str(i)].value
        player_position = ws["D" + str(i)].value
        player_name = translit(player_name, "ru", reversed=True)
        players[guest_team][player_number] = Player(player_number, 
                                                    swapper(player_name), 
                                                    player_position)
        
    
    periods = [[52, 76],
               [77, 98],
               [99, 119],
               [121, 121],
               [123, 133],
               ]
    goal_events = ['scored', 'score']
    injury_events = ['injury']   
    penalty_box_events = ['penalty-box']
    
    for period_number in xrange(len(periods)):
        period = periods[period_number]
        first_period_start = period[0]
        first_perion_end = period[1]
        action_flag = False
        if period_number == 4:
            penalties_scored_after_overtime = 0
            last_actor = ""
            last_team = ""
            attempts = 0
            
        for i in xrange(first_period_start + 1, first_perion_end + 1):
            event_number = ws["A" + str(i)].value
            event_time = ws["B" + str(i)].value
            event_action = ws["C" + str(i)].value
            event_team = ws["D" + str(i)].value
            event_player_number = ws["E" + str(i)].value
            event_place = ws["F" + str(i)].value
            event_result = ws["J" + str(i)].value
            
            team = teams[event_team]
            try:
                player_name = players[team][event_player_number].name
            except KeyError:
                player_name = None
            
            if period_number == 4:
                attempts += 1
            
            # Goal events.
            if event_action in goal_events or event_result in goal_events:
                action_flag = True
                if team_names[home_team] == event_team:
                    scores[home_team] += 1
                else:
                    scores[guest_team] += 1

                    if period_number == 3:
                        # Overtime
                        print generate_overtime_ending_phrase(player_name,
                                                              event_team)
                        continue
                        
                        
                    if period_number == 4:
                        # Penalties.
                        penalties_scored_after_overtime += 1
                        last_actor = player_name
                        last_team = event_team
                        continue                   
                    
                if is_first_goal_scored is False:
                    # First goal event.
                    if home_final_score + guest_final_score == 1:
                        # 1-0 or 0-1.
                        print generate_single_goal_phrase(player_name, 
                                                          event_time)
                    else:
                        actor_2 = ws["H" + str(i)].value
                        actor_3 = ws["I" + str(i)].value
                        print generate_first_goal_phrase(event_team, 
                                   player_name, 
                                   event_time, 
                                   players[team][actor_2].name, 
                                   players[team][actor_3].name)
                        
                    is_first_goal_scored = True
                else:
                    if scores[home_team] == home_final_score and \
                        scores[guest_team] == guest_final_score:
                        print generate_final_goal_phrase(team, 
                                                     player_name, 
                                                     event_time, 
                                                     period_number)

                    else:
                    # Common goal event.
                        print generate_goals_phrase(event_team, 
                                player_name, 
                                event_time, 
                                str(scores[home_team]), 
                                str(scores[guest_team]), 
                                team == home_team)                
                    continue
            
            # Penalty-box events.
            if event_action in penalty_box_events or \
                event_result in penalty_box_events:
                print generate_penalty_phrase(player_name, event_time)
                
            # Injury events.
            if event_action in injury_events or \
                event_result in injury_events:
                print generate_injury_phrase(player_name, event_time)
        if period_number <= 2:
            print generate_comment_after_period(action_flag, 
                                            period_number + 1, 
                                            scores[home_team], 
                                            scores[guest_team], 
                                            team_names[home_team], 
                                            team_names[guest_team])
    if period_number == 4:
        print generate_after_penalties_phrase(last_actor, last_team, 
                                    penalties_scored_after_overtime, 
                                    attempts) 
    
    print generate_end_phrase(home_final_score, guest_final_score)
